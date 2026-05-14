import os
import re
import json
import logging
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QFileDialog, QLabel, QLineEdit, QTableWidget, 
                             QTableWidgetItem, QProgressBar, QGroupBox, 
                             QSplitter, QTextEdit, QMessageBox, QCheckBox,
                             QDateEdit, QComboBox, QTabWidget)

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class InvoiceProcessor(QThread):
    """发票处理线程"""
    progress_updated = pyqtSignal(int)
    processing_complete = pyqtSignal(list)
    processing_error = pyqtSignal(str)
    
    def __init__(self, pdf_folder, excel_path):
        super().__init__()
        self.pdf_folder = pdf_folder
        self.excel_path = excel_path
        
    def extract_text_from_pdf(self, pdf_path):
        """从PDF提取文本"""
        try:
            logger.info(f"开始读取PDF: {pdf_path}")
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    text += page_text
                    logger.debug(f"第{i+1}页文本: {page_text[:200]}...")
                
                logger.info(f"PDF读取完成，总长度: {len(text)}字符")
                logger.info(f"完整文本内容:\n{text}")
                return text
        except Exception as e:
            logger.error(f"读取PDF失败: {str(e)}")
            return f"读取PDF失败: {str(e)}"
    
    def parse_invoice_data(self, text, filename):
        """解析发票数据 - 根据具体格式优化"""
        logger.info(f"开始解析发票数据: {filename}")
        logger.info(f"原始文本内容:\n{text}")
        
        data = {
            '发票号码': '提取失败',
            '发票日期': '提取失败',
            '发票类型': '提取失败',
            '购买方名称': '提取失败',
            '销售方名称': '提取失败',
            '项目名称': '提取失败',
            '金额': '提取失败',
            '税率': '提取失败',
            '税额': '提取失败',
            '价税合计': '提取失败',
            '备注': '增值'
        }
        
        # 发票号码 - 8-20位数字（放宽限制）
        logger.info("正在查找发票号码...")
        invoice_number_match = re.search(r'(?:发票号码|发票代码|制)[:：\s]*(\d{8,20})', text)
        if not invoice_number_match:
            invoice_number_match = re.search(r'\b(\d{8,20})\b', text)
        if not invoice_number_match:
            # 直接查找长数字序列
            invoice_number_match = re.search(r'(\d{16,20})', text)
        if invoice_number_match:
            data['发票号码'] = invoice_number_match.group(1)
            logger.info(f"找到发票号码: {data['发票号码']}")
        else:
            logger.warning("未找到发票号码")
        
        # 发票日期 - 支持多种格式
        logger.info("正在查找发票日期...")
        date_patterns = [
            r'(?:开票日期|日期)[:：\s]*(\d{4}[年\-/.]\d{1,2}[月\-/.]\d{1,2}[日]?)',
            r'(\d{4})年(\d{1,2})月(\d{1,2})日',
            r'(\d{4})-(\d{1,2})-(\d{1,2})',
            r'(\d{4})/(\d{1,2})/(\d{1,2})',
            r'(\d{4})\.(\d{1,2})\.(\d{1,2})'
        ]
        for pattern in date_patterns:
            date_match = re.search(pattern, text)
            if date_match:
                date_groups = date_match.groups()
                if len(date_groups) == 1:
                    date_str = date_groups[0]
                    date_str = re.sub(r'[年月日]', '', date_str)
                    date_str = date_str.replace('-', '').replace('/', '').replace('.', '')
                    if len(date_str) == 8:
                        data['发票日期'] = date_str
                        logger.info(f"找到发票日期: {data['发票日期']}")
                        break
                else:
                    year, month, day = date_groups
                    data['发票日期'] = f"{year}{int(month):02d}{int(day):02d}"
                    logger.info(f"找到发票日期: {data['发票日期']}")
                    break
        
        # 发票类型
        logger.info("正在判断发票类型...")
        if '专用发票' in text or '增值税专用发票' in text:
            data['发票类型'] = '专票'
            logger.info("识别为专用发票")
        elif '普通发票' in text or '增值税普通发票' in text:
            data['发票类型'] = '普票'
            logger.info("识别为普通发票")
        else:
            logger.warning("无法识别发票类型")
        
        # 购买方名称 - 第一个包含"有限公司"的公司
        logger.info("正在查找购买方名称...")
        company_matches = re.findall(r'[\u4e00-\u9fa5（）()]*有限公司', text)
        if company_matches:
            data['购买方名称'] = company_matches[0]
            logger.info(f"找到购买方名称: {data['购买方名称']}")
        
        # 销售方名称 - 第二个包含"有限公司"的公司
        logger.info("正在查找销售方名称...")
        if len(company_matches) >= 2:
            data['销售方名称'] = company_matches[1]
            logger.info(f"找到销售方名称: {data['销售方名称']}")
        elif len(company_matches) == 1:
            data['销售方名称'] = company_matches[0]
            logger.info(f"找到销售方名称(单个): {data['销售方名称']}")
        
        # 项目名称 - 提取*和*中间的中文字
        logger.info("正在查找项目名称...")
        project_match = re.search(r'\*([^*]+)\*', text)
        if project_match:
            data['项目名称'] = project_match.group(1).strip()
            logger.info(f"找到项目名称: {data['项目名称']}")
        else:
            # 备选方案
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if '项目名称' in line and i + 1 < len(lines):
                    next_line = lines[i + 1]
                    chinese_match = re.search(r'[*]*([^*\d]+)', next_line.strip())
                    if chinese_match:
                        data['项目名称'] = chinese_match.group(1).strip()
                        logger.info(f"找到项目名称(备选): {data['项目名称']}")
                        break
        
        # 金额 - 提取\xa后的三个数字中的中间值（按数值大小）
        logger.info("正在查找金额...")
        # 使用更宽松的正则表达式匹配所有可能的\xa5符号和数字格式
        xa_amounts = re.findall(r'[¥￥\xa5]([\d,.]+)', text)
        # 记录找到的所有\xa5后的值
        logger.info(f"找到的所有金额值: {xa_amounts}")
        if xa_amounts:
            amounts_float = [float(amt.replace(',', '')) for amt in xa_amounts]
            if len(amounts_float) >= 3:
                amounts_float_sorted = sorted(amounts_float)
                data['金额'] = str(amounts_float_sorted[1])  # 按数值大小排序后取中间值（第二个）
            elif len(amounts_float) >= 2:
                data['金额'] = str(amounts_float[0])  # 如果只有两个，取第一个
            else:
                data['金额'] = str(amounts_float[0])
            logger.info(f"找到金额: {data['金额']}")
        else:
            # 备选方案
            amounts = re.findall(r'[￥¥]([\d,.]+)', text)
            if not amounts:
                amounts = re.findall(r'[\d,]+\.\d{2}', text)
            if amounts:
                amounts_float = [float(amt.replace(',', '')) for amt in amounts]
                if len(amounts_float) >= 3:
                    data['金额'] = str(amounts_float[1])  # 中间值
                elif len(amounts_float) >= 1:
                    data['金额'] = str(amounts_float[0])
                logger.info(f"找到金额(备选): {data['金额']}")
        
        # 税率 - 百分比
        logger.info("正在查找税率...")
        tax_rate_match = re.search(r'(\d+(?:\.\d+)?)%', text)
        if tax_rate_match:
            rate = float(tax_rate_match.group(1)) / 100
            data['税率'] = f"{rate:.2f}"
            logger.info(f"找到税率: {data['税率']}")
        
        # 税额 - 提取xa后的最小数字
        logger.info("正在查找税额...")
        # 使用更宽松的正则表达式匹配所有可能的货币符号和数字格式
        tax_amounts = re.findall(r'[¥￥\xa5]([\d,.]+)', text)
        # 记录找到的所有货币符号后的值
        logger.info(f"找到的所有税额值: {tax_amounts}")
        if tax_amounts:
            amounts_float_tax = [float(amt.replace(',', '')) for amt in tax_amounts]
            data['税额'] = str(min(amounts_float_tax))  # 最小值
            logger.info(f"找到税额: {data['税额']}")
        elif amounts:
            amounts_float = [float(amt.replace(',', '')) for amt in amounts]
            if amounts_float:
                data['税额'] = str(min(amounts_float))
                logger.info(f"找到税额(备选): {data['税额']}")
        
        # 价税合计 - 严格提取中文大写金额并转换为数字
        logger.info("正在查找中文大写价税合计...")
        
        # 精确的中文大写金额匹配模式
        chinese_amount_patterns = [
            r'(?:价税合计|人民币|￥|¥|大写|合计)[:：\s]*([零壹贰叁肆伍陆柒捌玖拾佰仟万亿]+(?:圆|元)[零壹贰叁肆伍陆柒捌玖拾角分]*)',
            r'([零壹贰叁肆伍陆柒捌玖拾佰仟万亿]+(?:圆|元)[零壹贰叁肆伍陆柒捌玖拾角分]*)',
            r'(?:大写金额)[:：\s]*([零壹贰叁肆伍陆柒捌玖拾佰仟万亿]+(?:圆|元)[零壹贰叁肆伍陆柒捌玖拾角分]*)'
        ]
        
        def chinese_to_decimal(chinese_str):
            """将中文大写金额转换为数字"""
            chinese_str = chinese_str.strip()
            if not chinese_str:
                return 0.0
            
            # 中文数字映射
            chinese_num_map = {
                '零': 0, '壹': 1, '贰': 2, '叁': 3, '肆': 4, '伍': 5,
                '陆': 6, '柒': 7, '捌': 8, '玖': 9
            }
            
            # 单位映射
            unit_map = {
                '拾': 10, '佰': 100, '仟': 1000, '万': 10000, '亿': 100000000
            }
            
            # 分割整数和小数部分
            if '圆' in chinese_str:
                parts = chinese_str.split('圆')
            elif '元' in chinese_str:
                parts = chinese_str.split('元')
            else:
                parts = [chinese_str, '']
            
            integer_part = parts[0]
            decimal_part = parts[1] if len(parts) > 1 else ''
            
            # 转换整数部分
            result = 0
            temp = 0
            section = 0
            
            for char in integer_part:
                if char in chinese_num_map:
                    temp = chinese_num_map[char]
                elif char in unit_map:
                    unit = unit_map[char]
                    if unit >= 10000:  # 万、亿
                        result = (result + temp) * unit
                        section += result
                        result = 0
                        temp = 0
                    else:  # 拾、佰、仟
                        temp *= unit
                        result += temp
                        temp = 0
            
            result += section + temp
            
            # 转换小数部分（角分）
            decimal = 0.0
            if decimal_part:
                jiao = 0
                fen = 0
                
                # 提取角
                jiao_match = re.search(r'([零壹贰叁肆伍陆柒捌玖])角', decimal_part)
                if jiao_match:
                    jiao = chinese_num_map.get(jiao_match.group(1), 0)
                
                # 提取分
                fen_match = re.search(r'([零壹贰叁肆伍陆柒捌玖])分', decimal_part)
                if fen_match:
                    fen = chinese_num_map.get(fen_match.group(1), 0)
                
                decimal = jiao * 0.1 + fen * 0.01
            
            return result + decimal
        
        chinese_amount = None
        for pattern in chinese_amount_patterns:
            chinese_match = re.search(pattern, text)
            if chinese_match:
                chinese_amount = chinese_match.group(1)
                logger.info(f"找到中文大写金额: {chinese_amount}")
                break
        
        if chinese_amount:
            try:
                decimal_value = chinese_to_decimal(chinese_amount)
                data['价税合计'] = str(decimal_value)
                logger.info(f"中文金额转换结果: {chinese_amount} -> {decimal_value}")
            except Exception as e:
                logger.error(f"中文金额转换失败: {e}")
                # 转换失败时使用最大数字金额作为备选
                if amounts and amounts_float:
                    data['价税合计'] = str(amounts_float[-1])
                    logger.info(f"使用备选数字金额: {data['价税合计']}")
                else:
                    data['价税合计'] = '0.00'
        elif amounts and amounts_float:
            # 如果没有中文大写金额，才使用数字金额
            data['价税合计'] = str(amounts_float[-1])
            logger.info(f"未找到中文大写金额，使用数字金额: {data['价税合计']}")
        else:
            data['价税合计'] = '0.00'
        
        # 备注 - 根据内容提取
        logger.info("正在查找备注...")
        remark_keywords = ['机票', '酒店', '用车', '火车', '餐饮', '交通', '办公用品']
        for keyword in remark_keywords:
            if keyword in text:
                data['备注'] = keyword
                logger.info(f"找到备注: {data['备注']}")
                break
        else:
            data['备注'] = '增值'
            logger.info(f"使用默认备注: {data['备注']}")
        
        logger.info(f"解析完成，结果: {json.dumps(data, ensure_ascii=False, indent=2)}")
        return data
    
    def run(self):
        """执行发票处理"""
        try:
            logger.info(f"开始处理文件夹: {self.pdf_folder}")
            pdf_files = []
            for file in os.listdir(self.pdf_folder):
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(self.pdf_folder, file))
            
            logger.info(f"找到 {len(pdf_files)} 个PDF文件: {[os.path.basename(f) for f in pdf_files]}")
            
            if not pdf_files:
                self.processing_error.emit("未找到PDF文件")
                return
            
            results = []
            total_files = len(pdf_files)
            
            for i, pdf_file in enumerate(pdf_files):
                try:
                    filename = os.path.basename(pdf_file)
                    logger.info(f"正在处理第 {i+1}/{total_files} 个文件: {filename}")
                    text = self.extract_text_from_pdf(pdf_file)
                    invoice_data = self.parse_invoice_data(text, filename)
                    results.append(invoice_data)
                    
                    progress = int((i + 1) / total_files * 100)
                    self.progress_updated.emit(progress)
                    
                    logger.info(f"文件 {filename} 处理完成，结果: {json.dumps(invoice_data, ensure_ascii=False)}")
                    
                except Exception as e:
                    # 记录失败的文件
                    error_data = {
                        '发票号码': '提取失败',
                        '发票日期': '',
                        '发票类型': '',
                        '购买方名称': '',
                        '销售方名称': '',
                        '项目名称': '',
                        '金额': '',
                        '税率': '',
                        '税额': '',
                        '价税合计': '',
                        '备注': f'{filename} - 错误: {str(e)}'
                    }
                    results.append(error_data)
                    logger.error(f"处理文件 {filename} 失败: {str(e)}")
            
            logger.info(f"所有文件处理完成，共处理 {len(results)} 个文件")
            self.processing_complete.emit(results)
            
        except Exception as e:
            logger.error(f"处理过程发生错误: {str(e)}")
            self.processing_error.emit(str(e))


class CustomsExtractorWidget(QWidget):
    """海关提取子界面"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("海关信息提取")
        self.init_ui()
    
    def init_ui(self):
        """初始化海关提取界面"""
        layout = QVBoxLayout(self)
        
        # 文件选择区域
        file_group = QGroupBox("文件选择")
        file_layout = QHBoxLayout()
        
        self.folder_path = QLineEdit()
        self.folder_path.setPlaceholderText("请选择包含PDF海关单的文件夹...")
        
        browse_btn = QPushButton("📁 浏览")
        browse_btn.clicked.connect(self.browse_folder)
        
        file_layout.addWidget(QLabel("文件夹:"))
        file_layout.addWidget(self.folder_path)
        file_layout.addWidget(browse_btn)
        
        file_group.setLayout(file_layout)
        
        # 操作按钮区域
        button_layout = QHBoxLayout()
        
        self.process_btn = QPushButton("🚀 开始提取")
        self.process_btn.clicked.connect(self.start_processing)
        self.process_btn.setEnabled(False)
        
        self.export_btn = QPushButton("💾 导出Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)
        
        button_layout.addWidget(self.process_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        
        # 结果显示区域
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(4)
        self.result_table.setHorizontalHeaderLabels([
            "海关单号", "金额", "日期", "文件名"
        ])
        
        # 设置表格样式
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setStyleSheet("""
            QTableWidget {
                font-size: 16px;
                font-family: 'Segoe UI', 'Microsoft YaHei', '等线';
            }
            QTableWidget::item {
                padding: 8px;
            }
        """)
        
        # 日志显示区域
        self.log_display = QTextEdit()
        self.log_display.setMaximumHeight(150)
        self.log_display.setPlaceholderText("处理日志将显示在这里...")
        
        # 添加到主布局
        layout.addWidget(file_group)
        layout.addLayout(button_layout)
        layout.addWidget(self.progress)
        layout.addWidget(self.result_table)
        layout.addWidget(self.log_display)
        
        # 设置布局间距
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 初始化数据
        self.results = []
    
    def browse_folder(self):
        """浏览文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择海关PDF文件夹")
        if folder:
            self.folder_path.setText(folder)
            self.process_btn.setEnabled(True)
    
    def extract_customs_data(self, text, filename):
        """提取海关数据"""
        data = {
            '海关单号': '提取失败',
            '金额': '提取失败',
            '日期': '提取失败',
            '文件名': filename
        }
        
        # 提取海关单号 - 18位数字加后缀格式：18位纯数字 + -A或Lxx
        customs_number_match = re.search(r'(?:海关|单号)[:：\s]*(\d{18}-[A-Z]\d{0,2})', text)
        if not customs_number_match:
            customs_number_match = re.search(r'(\d{18}-[A-Z]\d{0,2})', text)
        
        if customs_number_match:
            data['海关单号'] = customs_number_match.group(1)
        else:
            # 尝试匹配18位纯数字
            pure_number_match = re.search(r'\b(\d{18})\b', text)
            if pure_number_match:
                data['海关单号'] = pure_number_match.group(1)
        
        # 提取金额 - 提取￥后的金额（包含小数点）
        amounts = re.findall(r'￥(\d+(?:,\d{3})*(?:\.\d{2})?)', text)
        if amounts:
            amounts_float = [float(amt.replace(',', '')) for amt in amounts]
            data['金额'] = str(max(amounts_float))
        
        # 提取填发日期 - 格式：填发日期后的yyyy年m月d日（支持1-2位月份和日期）
        logger.info(f"开始提取日期，PDF文本内容长度: {len(text)}")
        logger.info(f"PDF文本内容: {text[:500]}...")  # 打印前500字符
        
        # 修正后的正则：支持1-2位月份和日期
        fill_date_pattern = r'填发日期[：:\s]*(\d{4})年(\d{1,2})月(\d{1,2})日'
        date_match = re.search(fill_date_pattern, text)
        
        if date_match:
            year, month, day = date_match.groups()
            # 补零格式化月份和日期为两位数
            month = month.zfill(2)
            day = day.zfill(2)
            data['日期'] = f"{year}{month}{day}"
            logger.info(f"成功提取填发日期: {data['日期']}")
        else:
            logger.warning(f"未找到填发日期模式: {fill_date_pattern}")
            
            # 备选：查找任何yyyy年m月d日格式（支持1-2位）
            backup_pattern = r'(\d{4})年(\d{1,2})月(\d{1,2})日'
            backup_matches = re.findall(backup_pattern, text)
            logger.info(f"找到的所有日期格式: {backup_matches}")
            
            if backup_matches:
                year, month, day = backup_matches[0]  # 取第一个找到的日期
                # 补零格式化
                month = month.zfill(2)
                day = day.zfill(2)
                data['日期'] = f"{year}{month}{day}"
                logger.info(f"使用备选日期: {data['日期']}")
            else:
                logger.error("未找到任何日期格式")
                data['日期'] = "提取失败"
        
        return data
    
    def start_processing(self):
        """开始处理海关文件"""
        folder = self.folder_path.text()
        if not folder or not os.path.exists(folder):
            QMessageBox.warning(self, "警告", "请选择有效的文件夹")
            return
        
        self.process_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        
        # 清空之前的结果
        self.results.clear()
        self.result_table.setRowCount(0)
        
        try:
            pdf_files = []
            for file in os.listdir(folder):
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(folder, file))
            
            if not pdf_files:
                QMessageBox.information(self, "提示", "未找到PDF文件")
                return
            
            total_files = len(pdf_files)
            for i, pdf_file in enumerate(pdf_files):
                try:
                    filename = os.path.basename(pdf_file)
                    
                    # 读取PDF文本
                    with pdfplumber.open(pdf_file) as pdf:
                        text = ""
                        for page in pdf.pages:
                            page_text = page.extract_text() or ""
                            text += page_text
                    
                    customs_data = self.extract_customs_data(text, filename)
                    self.results.append(customs_data)
                    
                    # 更新表格
                    row = self.result_table.rowCount()
                    self.result_table.insertRow(row)
                    
                    self.result_table.setItem(row, 0, QTableWidgetItem(customs_data['海关单号']))
                    self.result_table.setItem(row, 1, QTableWidgetItem(customs_data['金额']))
                    self.result_table.setItem(row, 2, QTableWidgetItem(customs_data['日期']))
                    self.result_table.setItem(row, 3, QTableWidgetItem(customs_data['文件名']))
                    
                    # 更新进度
                    progress = int((i + 1) / total_files * 100)
                    self.progress.setValue(progress)
                    
                    # 更新日志
                    self.log_display.append(f"已处理: {filename}")
                    
                except Exception as e:
                    self.log_display.append(f"处理失败: {filename} - {str(e)}")
            
            self.export_btn.setEnabled(True)
            QMessageBox.information(self, "完成", f"处理完成，共处理 {len(self.results)} 个文件")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理失败: {str(e)}")
        finally:
            self.process_btn.setEnabled(True)
            self.progress.setVisible(False)
    
    def export_to_excel(self):
        """导出到Excel"""
        if not self.results:
            QMessageBox.warning(self, "警告", "没有数据可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", "海关信息.xlsx", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "海关信息"
                
                # 设置标题
                headers = ["海关单号", "金额", "日期", "文件名"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                
                # 填充数据
                for row, data in enumerate(self.results, 2):
                    ws.cell(row=row, column=1, value=data['海关单号'])
                    ws.cell(row=row, column=2, value=float(data['金额']) if data['金额'] != '提取失败' else 0)
                    ws.cell(row=row, column=3, value=data['日期'])
                    ws.cell(row=row, column=4, value=data['文件名'])
                
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(file_path)
                QMessageBox.information(self, "成功", f"已保存到: {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")


class CtripExtractorWidget(QWidget):
    """携程提取子界面"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("携程信息提取")
        self.init_ui()
    
    def init_ui(self):
        """初始化携程提取界面"""
        layout = QVBoxLayout(self)
        
        # 文件选择区域
        file_group = QGroupBox("文件选择")
        file_layout = QHBoxLayout()
        
        self.folder_path = QLineEdit()
        self.folder_path.setPlaceholderText("请选择包含PDF文件的文件夹...")
        
        browse_btn = QPushButton("📁 浏览")
        browse_btn.clicked.connect(self.browse_folder)
        
        file_layout.addWidget(QLabel("文件夹:"))
        file_layout.addWidget(self.folder_path)
        file_layout.addWidget(browse_btn)
        
        file_group.setLayout(file_layout)
        
        # 操作按钮区域
        button_layout = QHBoxLayout()
        
        self.process_btn = QPushButton("🚀 开始提取")
        self.process_btn.clicked.connect(self.start_processing)
        self.process_btn.setEnabled(False)
        
        self.export_btn = QPushButton("💾 导出Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)
        
        button_layout.addWidget(self.process_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        
        # 结果显示区域
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(5)
        self.result_table.setHorizontalHeaderLabels([
            "订单号", "金额", "日期", "旅客姓名", "文件名"
        ])
        
        # 设置表格样式
        self.result_table.setAlternatingRowColors(True)
        self.result_table.setStyleSheet("""
            QTableWidget {
                font-size: 16px;
                font-family: 'Segoe UI', 'Microsoft YaHei', '等线';
            }
            QTableWidget::item {
                padding: 8px;
            }
        """)
        
        # 日志显示区域
        self.log_display = QTextEdit()
        self.log_display.setMaximumHeight(150)
        self.log_display.setPlaceholderText("处理日志将显示在这里...")
        
        # 添加到主布局
        layout.addWidget(file_group)
        layout.addLayout(button_layout)
        layout.addWidget(self.progress)
        layout.addWidget(self.result_table)
        layout.addWidget(self.log_display)
        
        # 设置布局间距
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 初始化数据
        self.results = []
    
    def browse_folder(self):
        """浏览文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择PDF文件夹")
        if folder:
            self.folder_path.setText(folder)
            self.process_btn.setEnabled(True)
    
    def extract_ctrip_data(self, text, filename):
        """提取携程数据"""
        data = {
            '订单号': '提取失败',
            '金额': '提取失败',
            '日期': '提取失败',
            '旅客姓名': '提取失败',
            '文件名': filename
        }
        
        # 提取订单号 - 通常为10-12位数字
        order_patterns = [
            r'(?:订单号|订单编号|预订编号)[:：\s]*(\d{10,12})',
            r'(?:订单|预订)[:：\s]*(\d{10,12})',
            r'(\d{10,12})'
        ]
        
        for pattern in order_patterns:
            order_match = re.search(pattern, text)
            if order_match:
                data['订单号'] = order_match.group(1)
                break
        
        # 提取金额 - 寻找货币符号后的数字
        amount_patterns = [
            r'[¥￥\xa5]([\d,]+(?:\.\d{2})?)',
            r'(?:金额|费用|价格|总价)[:：\s]*([\d,]+(?:\.\d{2})?)',
            r'([\d,]+\.\d{2})'
        ]
        
        for pattern in amount_patterns:
            amount_matches = re.findall(pattern, text)
            if amount_matches:
                # 转换为浮点数并取最大值
                amounts = [float(amt.replace(',', '')) for amt in amount_matches]
                data['金额'] = str(max(amounts))
                break
        
        # 提取日期 - 多种日期格式
        date_patterns = [
            r'(?:出发日期|旅行日期|预订日期|日期)[:：\s]*(\d{4})[年\-/.](\d{1,2})[月\-/.](\d{1,2})[日]?',
            r'(\d{4})[年\-/.](\d{1,2})[月\-/.](\d{1,2})[日]?',
            r'(\d{4})-(\d{1,2})-(\d{1,2})',
            r'(\d{4})/(\d{1,2})/(\d{1,2})'
        ]
        
        for pattern in date_patterns:
            date_match = re.search(pattern, text)
            if date_match:
                year, month, day = date_match.groups()
                # 格式化为YYYYMMDD
                data['日期'] = f"{year}{int(month):02d}{int(day):02d}"
                break
        
        # 提取旅客姓名 - 查找中文姓名模式
        name_patterns = [
            r'(?:旅客|乘客|姓名|乘机人)[:：\s]*([^\n\r]{2,4})(?=[\s,，、])',
            r'(?:旅客|乘客|姓名|乘机人)[:：\s]*([^\n\r]{2,4})',
            r'([王李张刘陈杨黄赵吴周徐孙马朱胡郭何高林罗郑梁谢宋唐许韩冯邓曹彭曾萧田董袁潘于蒋蔡余杜叶程苏魏吕丁任沈姚卢姜崔钟谭陆汪范金石廖贾夏韦付方白邹孟熊秦邱江尹薛闫段雷侯龙史陶黎贺顾毛郝龚邵万钱严覃武戴莫孔向汤][^\s,，、]{1,2})'
        ]
        
        for pattern in name_patterns:
            name_match = re.search(pattern, text)
            if name_match:
                name = name_match.group(1).strip()
                # 过滤掉明显不是姓名的内容
                if len(name) >= 2 and len(name) <= 4 and not any(c in name for c in '0123456789@#$%^&*()'):
                    data['旅客姓名'] = name
                    break
        
        return data
    
    def start_processing(self):
        """开始处理携程文件"""
        folder = self.folder_path.text()
        if not folder or not os.path.exists(folder):
            QMessageBox.warning(self, "警告", "请选择有效的文件夹")
            return
        
        # 检查pdfplumber是否可用
        try:
            import pdfplumber
        except ImportError:
            QMessageBox.warning(self, "警告", "请先安装pdfplumber库！\n使用命令: pip install pdfplumber")
            return
        
        self.process_btn.setEnabled(False)
        self.export_btn.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        
        # 清空之前的结果
        self.results.clear()
        self.result_table.setRowCount(0)
        self.log_display.clear()
        
        try:
            pdf_files = []
            for file in os.listdir(folder):
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(folder, file))
            
            if not pdf_files:
                QMessageBox.information(self, "提示", "未找到PDF文件")
                return
            
            total_files = len(pdf_files)
            for i, pdf_file in enumerate(pdf_files):
                try:
                    filename = os.path.basename(pdf_file)
                    
                    # 读取PDF文本
                    try:
                        with pdfplumber.open(pdf_file) as pdf:
                            text = ""
                            for page in pdf.pages:
                                page_text = page.extract_text() or ""
                                text += page_text
                    except Exception as pdf_error:
                        self.log_display.append(f"PDF读取失败: {filename} - {str(pdf_error)}")
                        # 创建一个失败的记录
                        ctrip_data = {
                            '订单号': 'PDF读取失败',
                            '金额': 'PDF读取失败',
                            '日期': 'PDF读取失败',
                            '旅客姓名': 'PDF读取失败',
                            '文件名': filename
                        }
                        self.results.append(ctrip_data)
                        
                        # 更新表格
                        row = self.result_table.rowCount()
                        self.result_table.insertRow(row)
                        
                        self.result_table.setItem(row, 0, QTableWidgetItem(ctrip_data['订单号']))
                        self.result_table.setItem(row, 1, QTableWidgetItem(ctrip_data['金额']))
                        self.result_table.setItem(row, 2, QTableWidgetItem(ctrip_data['日期']))
                        self.result_table.setItem(row, 3, QTableWidgetItem(ctrip_data['旅客姓名']))
                        self.result_table.setItem(row, 4, QTableWidgetItem(ctrip_data['文件名']))
                        
                        # 更新进度
                        progress = int((i + 1) / total_files * 100)
                        self.progress.setValue(progress)
                        continue
                    
                    ctrip_data = self.extract_ctrip_data(text, filename)
                    self.results.append(ctrip_data)
                    
                    # 更新表格
                    row = self.result_table.rowCount()
                    self.result_table.insertRow(row)
                    
                    self.result_table.setItem(row, 0, QTableWidgetItem(ctrip_data['订单号']))
                    self.result_table.setItem(row, 1, QTableWidgetItem(ctrip_data['金额']))
                    self.result_table.setItem(row, 2, QTableWidgetItem(ctrip_data['日期']))
                    self.result_table.setItem(row, 3, QTableWidgetItem(ctrip_data['旅客姓名']))
                    self.result_table.setItem(row, 4, QTableWidgetItem(ctrip_data['文件名']))
                    
                    # 更新进度
                    progress = int((i + 1) / total_files * 100)
                    self.progress.setValue(progress)
                    
                    # 更新日志
                    self.log_display.append(f"已处理: {filename}")
                    
                except Exception as e:
                    self.log_display.append(f"处理失败: {filename} - {str(e)}")
            
            self.export_btn.setEnabled(True)
            success_count = len([r for r in self.results if r['订单号'] != 'PDF读取失败' and r['订单号'] != '提取失败'])
            total_count = len(self.results)
            QMessageBox.information(self, "完成", f"处理完成！\n成功：{success_count} 个\n失败：{total_count - success_count} 个")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"处理失败: {str(e)}")
        finally:
            self.process_btn.setEnabled(True)
            self.progress.setVisible(False)
    
    def export_to_excel(self):
        """导出到Excel"""
        if not self.results:
            QMessageBox.warning(self, "警告", "没有数据可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", "携程提取.xlsx", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "携程信息"
                
                # 设置标题
                headers = ["订单号", "金额", "日期", "旅客姓名", "文件名"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                
                # 填充数据
                for row, data in enumerate(self.results, 2):
                    ws.cell(row=row, column=1, value=data['订单号'])
                    ws.cell(row=row, column=2, value=float(data['金额']) if data['金额'] != '提取失败' else 0)
                    ws.cell(row=row, column=3, value=data['日期'])
                    ws.cell(row=row, column=4, value=data['旅客姓名'])
                    ws.cell(row=row, column=5, value=data['文件名'])
                
                # 调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(file_path)
                QMessageBox.information(self, "成功", f"已保存到: {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")


class InvoiceExtractorWidget(QWidget):
    """发票提取工具主界面"""
    
    def __init__(self):
        super().__init__()
        self.processor = None
        self.current_results = []
        self.init_ui()
        self.load_settings()
    
    def init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout(self)
        
        # 创建标签页
        self.tab_widget = QTabWidget()
        
        # 发票提取页面
        self.invoice_tab = QWidget()
        self.init_invoice_tab()
        
        # 海关提取页面
        self.customs_tab = CustomsExtractorWidget()
        
        # 携程提取页面
        self.ctrip_tab = CtripExtractorWidget()
        
        self.tab_widget.addTab(self.invoice_tab, "🧾 发票提取")
        self.tab_widget.addTab(self.customs_tab, "📋 海关提取")
        self.tab_widget.addTab(self.ctrip_tab, "✈️ 携程提取")
        
        layout.addWidget(self.tab_widget)
    
    def init_invoice_tab(self):
        """初始化发票提取标签页"""
        layout = QVBoxLayout(self.invoice_tab)
        
        # 创建主分割器
        splitter = QSplitter(Qt.Vertical)
        
        # 上半部分：控制面板
        control_panel = QWidget()
        control_layout = QVBoxLayout(control_panel)
        
        # PDF文件夹选择
        pdf_group = QGroupBox("PDF文件夹")
        pdf_layout = QHBoxLayout()
        self.pdf_path_input = QLineEdit()
        self.pdf_path_input.setPlaceholderText("选择包含PDF发票的文件夹...")
        pdf_browse_btn = QPushButton("浏览...")
        pdf_browse_btn.clicked.connect(self.browse_pdf_folder)
        pdf_layout.addWidget(self.pdf_path_input)
        pdf_layout.addWidget(pdf_browse_btn)
        pdf_group.setLayout(pdf_layout)
        control_layout.addWidget(pdf_group)
        
        # Excel文件选择
        excel_group = QGroupBox("Excel输出文件")
        excel_layout = QHBoxLayout()
        self.excel_path_input = QLineEdit()
        self.excel_path_input.setPlaceholderText("选择或创建Excel文件...")
        excel_browse_btn = QPushButton("浏览...")
        excel_browse_btn.clicked.connect(self.browse_excel_file)
        excel_layout.addWidget(self.excel_path_input)
        excel_layout.addWidget(excel_browse_btn)
        excel_group.setLayout(excel_layout)
        control_layout.addWidget(excel_group)
        
        # 处理控制区
        process_group = QGroupBox("处理控制")
        process_layout = QVBoxLayout()
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        process_layout.addWidget(self.progress_bar)
        
        # 处理按钮
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始提取")
        self.start_btn.clicked.connect(self.start_processing)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 12px 24px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        self.stop_btn = QPushButton("停止")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-weight: bold;
                padding: 12px 24px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
            QPushButton:pressed {
                background-color: #c62828;
            }
        """)
        
        self.save_btn = QPushButton("保存结果")
        self.save_btn.clicked.connect(self.save_results)
        self.save_btn.setEnabled(False)
        
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.stop_btn)
        btn_layout.addWidget(self.save_btn)
        btn_layout.addStretch()
        
        process_layout.addLayout(btn_layout)
        process_group.setLayout(process_layout)
        control_layout.addWidget(process_group)
        
        # 统计信息
        self.stats_label = QLabel("准备就绪")
        self.stats_label.setStyleSheet("font-weight: bold; color: #666;")
        control_layout.addWidget(self.stats_label)
        
        control_layout.addStretch()
        
        # 下半部分：结果显示
        result_widget = QWidget()
        result_layout = QVBoxLayout(result_widget)
        
        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(12)
        self.result_table.setHorizontalHeaderLabels([
            "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
            "项目名称", "金额", "税率", "税额", "价税合计", "备注", "文件名"
        ])
        
        # 设置表格样式
        self.result_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #ddd;
                font-weight: bold;
            }
        """)
        
        # 设置列宽
        self.result_table.setColumnWidth(0, 150)  # 发票号码
        self.result_table.setColumnWidth(1, 100)  # 发票日期
        self.result_table.setColumnWidth(2, 80)   # 发票类型
        self.result_table.setColumnWidth(3, 200)  # 购买方名称
        self.result_table.setColumnWidth(4, 200)  # 销售方名称
        self.result_table.setColumnWidth(5, 200)  # 项目名称
        self.result_table.setColumnWidth(6, 100)  # 金额
        self.result_table.setColumnWidth(7, 80)   # 税率
        self.result_table.setColumnWidth(8, 100)  # 税额
        self.result_table.setColumnWidth(9, 100)  # 价税合计
        self.result_table.setColumnWidth(10, 200) # 备注
        self.result_table.setColumnWidth(11, 150) # 文件名
        
        result_layout.addWidget(self.result_table)
        
        # 添加到分割器
        splitter.addWidget(control_panel)
        splitter.addWidget(result_widget)
        
        # 设置分割器比例
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        
        layout.addWidget(splitter)
    
    def browse_pdf_folder(self):
        """浏览PDF文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择PDF文件夹")
        if folder:
            self.pdf_path_input.setText(folder)
            self.save_settings()
    
    def browse_excel_file(self):
        """浏览Excel文件"""
        file, _ = QFileDialog.getSaveFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx);;所有文件 (*.*)"
        )
        if file:
            if not file.endswith('.xlsx'):
                file += '.xlsx'
            self.excel_path_input.setText(file)
            self.save_settings()
    
    def start_processing(self):
        """开始处理"""
        pdf_folder = self.pdf_path_input.text().strip()
        excel_path = self.excel_path_input.text().strip()
        
        if not pdf_folder or not os.path.exists(pdf_folder):
            QMessageBox.warning(self, "警告", "请选择有效的PDF文件夹！")
            return
        
        if not excel_path:
            QMessageBox.warning(self, "警告", "请选择Excel输出文件！")
            return
        
        # 检查pdfplumber是否可用
        try:
            import pdfplumber
        except ImportError:
            QMessageBox.warning(self, "警告", "请先安装pdfplumber库！")
            return
        
        # 设置界面状态
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.save_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        
        # 创建并启动处理线程
        self.processor = InvoiceProcessor(pdf_folder, excel_path)
        self.processor.progress_updated.connect(self.update_progress)
        self.processor.processing_complete.connect(self.processing_finished)
        self.processor.processing_error.connect(self.processing_error)
        self.processor.start()
    
    def stop_processing(self):
        """停止处理"""
        if self.processor and self.processor.isRunning():
            self.processor.terminate()
            self.processor.wait()
        
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.stats_label.setText("处理已停止")
    
    def update_progress(self, value):
        """更新进度"""
        self.progress_bar.setValue(value)
    
    def processing_finished(self, results):
        """处理完成"""
        self.current_results = results
        self.display_results(results)
        
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.save_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        
        success_count = len([r for r in results if r['发票号码'] != '提取失败'])
        total_count = len(results)
        self.stats_label.setText(f"处理完成：{success_count}/{total_count} 个文件成功")
        
        QMessageBox.information(self, "完成", f"处理完成！\n成功：{success_count} 个\n失败：{total_count - success_count} 个")
    
    def processing_error(self, error_msg):
        """处理错误"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        
        QMessageBox.critical(self, "错误", f"处理失败：\n{error_msg}")
        self.stats_label.setText(f"处理失败：{error_msg}")
    
    def display_results(self, results):
        """显示结果"""
        self.result_table.setRowCount(0)
        
        for row_idx, result in enumerate(results):
            self.result_table.insertRow(row_idx)
            
            for col_idx, field in enumerate([
                '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                '项目名称', '金额', '税率', '税额', '价税合计', '备注'
            ]):
                value = str(result.get(field, ''))
                item = QTableWidgetItem(value)
                
                # 标记错误行
                if value == '提取失败':
                    item.setBackground(Qt.red)
                    item.setForeground(Qt.white)
                
                self.result_table.setItem(row_idx, col_idx, item)
    
    def save_results(self):
        """保存结果到Excel"""
        if not self.current_results:
            QMessageBox.warning(self, "警告", "没有可保存的结果！")
            return
        
        excel_path = self.excel_path_input.text().strip()
        if not excel_path:
            return
        
        try:
            # 创建或加载Excel文件
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
                # 清除旧数据（保留表头）
                ws.delete_rows(2, ws.max_row)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "发票数据"
                
                # 写入表头
                headers = [
                    "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
                    "项目名称", "金额", "税率", "税额", "价税合计", "备注"
                ]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            
            # 写入数据
            for row_idx, result in enumerate(self.current_results, 2):
                for col_idx, field in enumerate([
                    '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                    '项目名称', '金额', '税率', '税额', '价税合计', '备注'
                ], 1):
                    value = result.get(field, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            QMessageBox.information(self, "成功", f"结果已保存到：\n{excel_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：\n{str(e)}")
    
    def load_settings(self):
        """加载设置"""
        settings_file = os.path.join(os.path.dirname(__file__), '..', 'toolbox_config.json')
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
                    if 'invoice_extractor' in settings:
                        config = settings['invoice_extractor']
                        self.pdf_path_input.setText(config.get('pdf_folder', ''))
                        self.excel_path_input.setText(config.get('excel_file', ''))
            except:
                pass
    
    def save_settings(self):
        """保存设置"""
        settings_file = os.path.join(os.path.dirname(__file__), '..', 'toolbox_config.json')
        settings = {}
        
        # 加载现有设置
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            except:
                pass
        
        # 更新发票提取器设置
        settings['invoice_extractor'] = {
            'pdf_folder': self.pdf_path_input.text(),
            'excel_file': self.excel_path_input.text()
        }
        
        # 保存设置
        try:
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=2, ensure_ascii=False)
        except:
            pass