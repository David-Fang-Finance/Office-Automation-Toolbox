import os
import re
import json
import base64
import logging
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QFileDialog, QLabel, QLineEdit, QTableWidget, 
                             QTableWidgetItem, QProgressBar, QGroupBox, 
                             QSplitter, QTextEdit, QMessageBox, QCheckBox,
                             QDateEdit, QComboBox, QTabWidget)

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class OCRInvoiceProcessor(QThread):
    """OCR发票处理线程"""
    progress_updated = pyqtSignal(int)
    processing_complete = pyqtSignal(list)
    processing_error = pyqtSignal(str)
    
    def __init__(self, pdf_folder, excel_path, api_key, secret_key):
        super().__init__()
        self.pdf_folder = pdf_folder
        self.excel_path = excel_path
        self.api_key = api_key
        self.secret_key = secret_key
        self.access_token = None
        
    def get_access_token(self):
        """获取百度API访问令牌"""
        try:
            url = 'https://aip.baidubce.com/oauth/2.0/token'
            params = {
                'grant_type': 'client_credentials',
                'client_id': self.api_key,
                'client_secret': self.secret_key
            }
            
            # 构建POST请求
            data = urlencode(params).encode('utf-8')
            req = Request(url, data=data)
            req.add_header('Content-Type', 'application/x-www-form-urlencoded')
            
            # 发送请求
            response = urlopen(req, timeout=30)
            result = response.read().decode('utf-8')
            
            # 解析JSON响应
            import json
            token_data = json.loads(result)
            
            if 'access_token' in token_data:
                self.access_token = token_data['access_token']
                logger.info("成功获取访问令牌")
                return True
            else:
                logger.error(f"获取访问令牌失败: {token_data}")
                return False
                
        except Exception as e:
            logger.error(f"获取访问令牌异常: {str(e)}")
            return False
    
    def extract_text_from_pdf(self, pdf_path):
        """使用OCR API从PDF提取文本"""
        try:
            logger.info(f"开始OCR处理PDF: {pdf_path}")
            
            # 获取访问令牌
            if not self.access_token:
                if not self.get_access_token():
                    return "获取访问令牌失败"
            
            # 读取PDF文件并转换为base64
            with open(pdf_path, 'rb') as f:
                pdf_content = f.read()
            
            # 将PDF转换为base64字符串
            pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
            
            # 调用百度OCR API - 使用高精度文字识别API
            url = f'https://aip.baidubce.com/rest/2.0/ocr/v1/accurate_basic?access_token={self.access_token}'
            
            params = {
                'pdf_file': pdf_base64,
                'detect_direction': 'true',
                'probability': 'false'
            }
            
            # 构建POST请求
            data = urlencode(params).encode('utf-8')
            req = Request(url, data=data)
            req.add_header('Content-Type', 'application/x-www-form-urlencoded')
            
            # 发送请求
            response = urlopen(req, timeout=60)
            result = response.read().decode('utf-8')
            
            # 解析OCR结果
            import json
            ocr_result = json.loads(result)
            
            if 'words_result' in ocr_result:
                # 提取所有文本
                text = ''
                for item in ocr_result['words_result']:
                    if 'words' in item:
                        text += item['words'] + '\n'
                
                logger.info(f"OCR处理完成，总长度: {len(text)}字符")
                logger.info(f"完整文本内容:\n{text}")
                return text
            else:
                logger.error(f"OCR处理失败: {ocr_result}")
                return f"OCR处理失败: {ocr_result}"
                
        except Exception as e:
            logger.error(f"OCR处理异常: {str(e)}")
            return f"OCR处理异常: {str(e)}"
    
    def parse_invoice_data(self, text, filename):
        """解析发票数据 - 使用与原来相同的解析逻辑"""
        logger.info(f"开始解析发票数据: {filename}")
        logger.info(f"原始文本内容:\n{text}")
        
        data = {
            '发票号码': '提取失败',
            '发票日期': '提取失败',
            '发票类型': '提取失败',
            '购买方名称': '提取失败',
            '销售方名称': '提取失败',
            '项目名称': '提取失败',
            '金额': '提取失败',
            '税率': '提取失败',
            '税额': '提取失败',
            '价税合计': '提取失败',
            '备注': '增值'
        }
        
        # 发票号码 - 8-20位数字（放宽限制）
        logger.info("正在查找发票号码...")
        invoice_number_match = re.search(r'(?:发票号码|发票代码|制)[:：\s]*(\d{8,20})', text)
        if not invoice_number_match:
            invoice_number_match = re.search(r'\b(\d{8,20})\b', text)
        if not invoice_number_match:
            # 直接查找长数字序列
            invoice_number_match = re.search(r'(\d{16,20})', text)
        if invoice_number_match:
            data['发票号码'] = invoice_number_match.group(1)
            logger.info(f"找到发票号码: {data['发票号码']}")
        else:
            logger.warning("未找到发票号码")
        
        # 发票日期 - 支持多种格式
        logger.info("正在查找发票日期...")
        date_patterns = [
            r'(?:开票日期|日期)[:：\s]*(\d{4}[年\-/.]\d{1,2}[月\-/.]\d{1,2}[日]?)',
            r'(\d{4})年(\d{1,2})月(\d{1,2})日',
            r'(\d{4})-(\d{1,2})-(\d{1,2})',
            r'(\d{4})/(\d{1,2})/(\d{1,2})',
            r'(\d{4})\.(\d{1,2})\.(\d{1,2})'
        ]
        for pattern in date_patterns:
            date_match = re.search(pattern, text)
            if date_match:
                date_groups = date_match.groups()
                if len(date_groups) == 1:
                    date_str = date_groups[0]
                    date_str = re.sub(r'[年月日]', '', date_str)
                    date_str = date_str.replace('-', '').replace('/', '').replace('.', '')
                    if len(date_str) == 8:
                        data['发票日期'] = date_str
                        logger.info(f"找到发票日期: {data['发票日期']}")
                        break
                else:
                    year, month, day = date_groups
                    data['发票日期'] = f"{year}{int(month):02d}{int(day):02d}"
                    logger.info(f"找到发票日期: {data['发票日期']}")
                    break
        
        # 发票类型
        logger.info("正在判断发票类型...")
        if '专用发票' in text or '增值税专用发票' in text:
            data['发票类型'] = '专票'
            logger.info("识别为专用发票")
        elif '普通发票' in text or '增值税普通发票' in text:
            data['发票类型'] = '普票'
            logger.info("识别为普通发票")
        else:
            logger.warning("无法识别发票类型")
        
        # 购买方名称 - 第一个包含"有限公司"的公司
        logger.info("正在查找购买方名称...")
        company_matches = re.findall(r'[\u4e00-\u9fa5（）()]*有限公司', text)
        if company_matches:
            data['购买方名称'] = company_matches[0]
            logger.info(f"找到购买方名称: {data['购买方名称']}")
        
        # 销售方名称 - 第二个包含"有限公司"的公司
        logger.info("正在查找销售方名称...")
        if len(company_matches) >= 2:
            data['销售方名称'] = company_matches[1]
            logger.info(f"找到销售方名称: {data['销售方名称']}")
        elif len(company_matches) == 1:
            data['销售方名称'] = company_matches[0]
            logger.info(f"找到销售方名称(单个): {data['销售方名称']}")
        
        # 项目名称 - 提取*和*中间的中文字
        logger.info("正在查找项目名称...")
        project_match = re.search(r'\*([^*]+)\*', text)
        if project_match:
            data['项目名称'] = project_match.group(1).strip()
            logger.info(f"找到项目名称: {data['项目名称']}")
        else:
            # 备选方案
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if '项目名称' in line and i + 1 < len(lines):
                    next_line = lines[i + 1]
                    chinese_match = re.search(r'[*]*([^*\d]+)', next_line.strip())
                    if chinese_match:
                        data['项目名称'] = chinese_match.group(1).strip()
                        logger.info(f"找到项目名称(备选): {data['项目名称']}")
                        break
        
        # 金额 - 提取\xa后的三个数字中的中间值（按数值大小）
        logger.info("正在查找金额...")
        # 使用更宽松的正则表达式匹配所有可能的\xa5符号和数字格式
        xa_amounts = re.findall(r'[¥￥\xa5]([\d,.]+)', text)
        # 记录找到的所有\xa5后的值
        logger.info(f"找到的所有金额值: {xa_amounts}")
        if xa_amounts:
            amounts_float = [float(amt.replace(',', '')) for amt in xa_amounts]
            if len(amounts_float) >= 3:
                amounts_float_sorted = sorted(amounts_float)
                data['金额'] = str(amounts_float_sorted[1])  # 按数值大小排序后取中间值（第二个）
            elif len(amounts_float) >= 2:
                data['金额'] = str(amounts_float[0])  # 如果只有两个，取第一个
            else:
                data['金额'] = str(amounts_float[0])
            logger.info(f"找到金额: {data['金额']}")
        else:
            # 备选方案
            amounts = re.findall(r'[￥¥]([\d,.]+)', text)
            if not amounts:
                amounts = re.findall(r'[\d,]+\.\d{2}', text)
            if amounts:
                amounts_float = [float(amt.replace(',', '')) for amt in amounts]
                if len(amounts_float) >= 3:
                    data['金额'] = str(amounts_float[1])  # 中间值
                elif len(amounts_float) >= 1:
                    data['金额'] = str(amounts_float[0])
                logger.info(f"找到金额(备选): {data['金额']}")
        
        # 税率 - 百分比
        logger.info("正在查找税率...")
        tax_rate_match = re.search(r'(\d+(?:\.\d+)?)%', text)
        if tax_rate_match:
            rate = float(tax_rate_match.group(1)) / 100
            data['税率'] = f"{rate:.2f}"
            logger.info(f"找到税率: {data['税率']}")
        
        # 税额 - 提取xa后的最小数字
        logger.info("正在查找税额...")
        # 使用更宽松的正则表达式匹配所有可能的货币符号和数字格式
        tax_amounts = re.findall(r'[¥￥\xa5]([\d,.]+)', text)
        # 记录找到的所有货币符号后的值
        logger.info(f"找到的所有税额值: {tax_amounts}")
        if tax_amounts:
            amounts_float = [float(amt.replace(',', '')) for amt in tax_amounts]
            if len(amounts_float) >= 2:
                # 税额通常是较小的那个
                data['税额'] = str(min(amounts_float))
            else:
                data['税额'] = str(amounts_float[0])
            logger.info(f"找到税额: {data['税额']}")
        
        # 价税合计 - 总金额
        logger.info("正在查找价税合计...")
        if data['金额'] != '提取失败' and data['税额'] != '提取失败':
            try:
                amount = float(data['金额'])
                tax = float(data['税额'])
                total = amount + tax
                data['价税合计'] = str(total)
                logger.info(f"计算价税合计: {data['价税合计']}")
            except:
                logger.warning("计算价税合计失败")
        
        # 备注 - 添加文件名信息
        data['备注'] = f"{filename} - OCR提取"
        
        return data
    
    def run(self):
        """主处理逻辑"""
        try:
            logger.info("开始处理发票提取任务")
            results = []
            
            # 获取PDF文件列表
            pdf_files = []
            for file in os.listdir(self.pdf_folder):
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(self.pdf_folder, file))
            
            total_files = len(pdf_files)
            logger.info(f"找到 {total_files} 个PDF文件")
            
            if total_files == 0:
                self.processing_error.emit("未找到PDF文件")
                return
            
            # 处理每个PDF文件
            for i, pdf_file in enumerate(pdf_files):
                try:
                    logger.info(f"处理文件 {i+1}/{total_files}: {os.path.basename(pdf_file)}")
                    
                    # 提取文本
                    text = self.extract_text_from_pdf(pdf_file)
                    
                    # 解析发票数据
                    invoice_data = self.parse_invoice_data(text, os.path.basename(pdf_file))
                    results.append(invoice_data)
                    
                    # 更新进度
                    progress = int((i + 1) / total_files * 100)
                    self.progress_updated.emit(progress)
                    
                except Exception as e:
                    logger.error(f"处理文件 {pdf_file} 失败: {str(e)}")
                    # 添加失败记录
                    results.append({
                        '发票号码': '提取失败',
                        '发票日期': '提取失败',
                        '发票类型': '提取失败',
                        '购买方名称': '提取失败',
                        '销售方名称': '提取失败',
                        '项目名称': '提取失败',
                        '金额': '提取失败',
                        '税率': '提取失败',
                        '税额': '提取失败',
                        '价税合计': '提取失败',
                        '备注': f"{os.path.basename(pdf_file)} - 处理失败"
                    })
            
            logger.info(f"处理完成，共提取 {len(results)} 个发票")
            self.processing_complete.emit(results)
            
        except Exception as e:
            logger.error(f"处理过程异常: {str(e)}")
            self.processing_error.emit(str(e))


class OCRInvoiceExtractorWidget(QWidget):
    """OCR发票提取工具主界面"""
    
    def __init__(self):
        super().__init__()
        self.processor = None
        self.current_results = []
        self.init_ui()
        self.load_settings()
    
    def init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout(self)
        
        # 创建分割器
        splitter = QSplitter(Qt.Vertical)
        
        # 上半部分：控制面板
        control_panel = QWidget()
        control_layout = QVBoxLayout(control_panel)
        
        # API Key输入
        api_group = QGroupBox("百度OCR API配置")
        api_layout = QVBoxLayout()
        
        # API Key
        api_key_layout = QHBoxLayout()
        api_key_layout.addWidget(QLabel("API Key:"))
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("请输入百度OCR API Key...")
        api_key_layout.addWidget(self.api_key_input)
        api_layout.addLayout(api_key_layout)
        
        # Secret Key
        secret_key_layout = QHBoxLayout()
        secret_key_layout.addWidget(QLabel("Secret Key:"))
        self.secret_key_input = QLineEdit()
        self.secret_key_input.setPlaceholderText("请输入百度OCR Secret Key...")
        secret_key_layout.addWidget(self.secret_key_input)
        api_layout.addLayout(secret_key_layout)
        
        api_group.setLayout(api_layout)
        control_layout.addWidget(api_group)
        
        # PDF文件夹选择
        pdf_group = QGroupBox("PDF文件夹")
        pdf_layout = QHBoxLayout()
        self.pdf_path_input = QLineEdit()
        self.pdf_path_input.setPlaceholderText("选择包含PDF发票的文件夹...")
        pdf_browse_btn = QPushButton("浏览...")
        pdf_browse_btn.clicked.connect(self.browse_pdf_folder)
        pdf_layout.addWidget(self.pdf_path_input)
        pdf_layout.addWidget(pdf_browse_btn)
        pdf_group.setLayout(pdf_layout)
        control_layout.addWidget(pdf_group)
        
        # Excel文件选择
        excel_group = QGroupBox("Excel输出文件")
        excel_layout = QHBoxLayout()
        self.excel_path_input = QLineEdit()
        self.excel_path_input.setPlaceholderText("选择或创建Excel文件...")
        excel_browse_btn = QPushButton("浏览...")
        excel_browse_btn.clicked.connect(self.browse_excel_file)
        excel_layout.addWidget(self.excel_path_input)
        excel_layout.addWidget(excel_browse_btn)
        excel_group.setLayout(excel_layout)
        control_layout.addWidget(excel_group)
        
        # 处理控制区
        process_group = QGroupBox("处理控制")
        process_layout = QVBoxLayout()
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        process_layout.addWidget(self.progress_bar)
        
        # 处理按钮
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始提取")
        self.start_btn.clicked.connect(self.start_processing)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 12px 24px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        self.stop_btn = QPushButton("停止")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-weight: bold;
                padding: 12px 24px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
            QPushButton:pressed {
                background-color: #c62828;
            }
        """)
        
        self.save_btn = QPushButton("保存结果")
        self.save_btn.clicked.connect(self.save_results)
        self.save_btn.setEnabled(False)
        
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.stop_btn)
        btn_layout.addWidget(self.save_btn)
        btn_layout.addStretch()
        
        process_layout.addLayout(btn_layout)
        process_group.setLayout(process_layout)
        control_layout.addWidget(process_group)
        
        # 统计信息
        self.stats_label = QLabel("准备就绪")
        self.stats_label.setStyleSheet("font-weight: bold; color: #666;")
        control_layout.addWidget(self.stats_label)
        
        control_layout.addStretch()
        
        # 下半部分：结果显示
        result_widget = QWidget()
        result_layout = QVBoxLayout(result_widget)
        
        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(12)
        self.result_table.setHorizontalHeaderLabels([
            "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
            "项目名称", "金额", "税率", "税额", "价税合计", "备注", "文件名"
        ])
        
        # 设置表格样式
        self.result_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #ddd;
                font-weight: bold;
            }
        """)
        
        # 设置列宽
        self.result_table.setColumnWidth(0, 150)  # 发票号码
        self.result_table.setColumnWidth(1, 100)  # 发票日期
        self.result_table.setColumnWidth(2, 80)   # 发票类型
        self.result_table.setColumnWidth(3, 200)  # 购买方名称
        self.result_table.setColumnWidth(4, 200)  # 销售方名称
        self.result_table.setColumnWidth(5, 200)  # 项目名称
        self.result_table.setColumnWidth(6, 100)  # 金额
        self.result_table.setColumnWidth(7, 80)   # 税率
        self.result_table.setColumnWidth(8, 100)  # 税额
        self.result_table.setColumnWidth(9, 100)  # 价税合计
        self.result_table.setColumnWidth(10, 200) # 备注
        self.result_table.setColumnWidth(11, 150) # 文件名
        
        result_layout.addWidget(self.result_table)
        
        # 添加到分割器
        splitter.addWidget(control_panel)
        splitter.addWidget(result_widget)
        
        # 设置分割器比例
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        
        layout.addWidget(splitter)
    
    def browse_pdf_folder(self):
        """浏览PDF文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择PDF文件夹", self.pdf_path_input.text())
        if folder:
            self.pdf_path_input.setText(folder)
    
    def browse_excel_file(self):
        """浏览Excel文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "选择Excel文件", self.excel_path_input.text(), "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            if not file_path.endswith(('.xlsx', '.xls')):
                file_path += '.xlsx'
            self.excel_path_input.setText(file_path)
    
    def start_processing(self):
        """开始处理"""
        pdf_folder = self.pdf_path_input.text()
        excel_path = self.excel_path_input.text()
        api_key = self.api_key_input.text()
        secret_key = self.secret_key_input.text()
        
        if not pdf_folder:
            QMessageBox.warning(self, "警告", "请选择PDF文件夹")
            return
        
        if not excel_path:
            QMessageBox.warning(self, "警告", "请选择Excel输出文件")
            return
        
        if not api_key:
            QMessageBox.warning(self, "警告", "请输入API Key")
            return
        
        if not secret_key:
            QMessageBox.warning(self, "警告", "请输入Secret Key")
            return
        
        if not os.path.exists(pdf_folder):
            QMessageBox.warning(self, "警告", "PDF文件夹不存在")
            return
        
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.save_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.stats_label.setText("正在处理...")
        
        self.processor = OCRInvoiceProcessor(pdf_folder, excel_path, api_key, secret_key)
        self.processor.progress_updated.connect(self.update_progress)
        self.processor.processing_complete.connect(self.processing_finished)
        self.processor.processing_error.connect(self.processing_error)
        self.processor.start()
    
    def stop_processing(self):
        """停止处理"""
        if self.processor and self.processor.isRunning():
            self.processor.terminate()
            self.processor.wait()
        
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.stats_label.setText("处理已停止")
    
    def update_progress(self, value):
        """更新进度"""
        self.progress_bar.setValue(value)
    
    def processing_finished(self, results):
        """处理完成"""
        self.current_results = results
        self.display_results(results)
        
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.save_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        
        # 保存设置
        self.save_settings()
        
        success_count = len([r for r in results if r['发票号码'] != '提取失败'])
        total_count = len(results)
        self.stats_label.setText(f"处理完成：{success_count}/{total_count} 个文件成功")
        
        QMessageBox.information(self, "完成", f"处理完成！\n成功：{success_count} 个\n失败：{total_count - success_count} 个")
    
    def processing_error(self, error_msg):
        """处理错误"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_bar.setVisible(False)
        
        QMessageBox.critical(self, "错误", f"处理失败：\n{error_msg}")
        self.stats_label.setText(f"处理失败：{error_msg}")
    
    def display_results(self, results):
        """显示结果"""
        self.result_table.setRowCount(0)
        
        for row_idx, result in enumerate(results):
            self.result_table.insertRow(row_idx)
            
            for col_idx, field in enumerate([
                '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                '项目名称', '金额', '税率', '税额', '价税合计', '备注'
            ]):
                value = str(result.get(field, ''))
                item = QTableWidgetItem(value)
                
                # 标记错误行
                if value == '提取失败':
                    item.setBackground(Qt.red)
                    item.setForeground(Qt.white)
                
                self.result_table.setItem(row_idx, col_idx, item)
            
            # 添加文件名列
            filename = os.path.basename(result.get('备注', '').split(' - ')[0] if ' - ' in result.get('备注', '') else '')
            filename_item = QTableWidgetItem(filename)
            self.result_table.setItem(row_idx, 11, filename_item)
    
    def save_results(self):
        """保存结果到Excel"""
        if not self.current_results:
            QMessageBox.warning(self, "警告", "没有可保存的结果")
            return
        
        excel_path = self.excel_path_input.text()
        if not excel_path:
            QMessageBox.warning(self, "警告", "请选择Excel输出文件")
            return
        
        try:
            # 创建或加载Excel文件
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
                # 清空现有数据（保留标题）
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.value = None
            else:
                wb = Workbook()
                ws = wb.active
                # 添加标题
                headers = [
                    '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                    '项目名称', '金额', '税率', '税额', '价税合计', '备注'
                ]
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="B3F6B0", end_color="B3F6B0", fill_type="solid")
            
            # 写入数据
            for row_idx, result in enumerate(self.current_results, 2):
                for col_idx, field in enumerate([
                    '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                    '项目名称', '金额', '税率', '税额', '价税合计', '备注'
                ], 1):
                    value = result.get(field, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            QMessageBox.information(self, "成功", f"结果已保存到：\n{excel_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败：\n{str(e)}")
    
    def load_settings(self):
        """加载设置"""
        settings_file = os.path.join(os.path.dirname(__file__), '..', 'toolbox_config.json')
        settings = {}
        
        # 加载现有设置
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            except:
                pass
        
        # 加载OCR发票提取器设置
        ocr_settings = settings.get('ocr_invoice_extractor', {})
        self.pdf_path_input.setText(ocr_settings.get('pdf_folder', ''))
        self.excel_path_input.setText(ocr_settings.get('excel_file', ''))
        self.api_key_input.setText(ocr_settings.get('api_key', ''))
        self.secret_key_input.setText(ocr_settings.get('secret_key', ''))
    
    def save_settings(self):
        """保存设置"""
        settings_file = os.path.join(os.path.dirname(__file__), '..', 'toolbox_config.json')
        settings = {}
        
        # 加载现有设置
        if os.path.exists(settings_file):
            try:
                with open(settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            except:
                pass
        
        # 更新OCR发票提取器设置
        settings['ocr_invoice_extractor'] = {
            'pdf_folder': self.pdf_path_input.text(),
            'excel_file': self.excel_path_input.text(),
            'api_key': self.api_key_input.text(),
            'secret_key': self.secret_key_input.text()
        }
        
        # 保存设置
        try:
            with open(settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, indent=2, ensure_ascii=False)
        except:
            pass