import os
import re
import json
import logging
import time
from datetime import datetime
from pathlib import Path

from google import genai
from google.genai import types
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                             QFileDialog, QLabel, QLineEdit, QTableWidget, 
                             QTableWidgetItem, QProgressBar, QGroupBox, 
                             QSplitter, QMessageBox, QTextEdit)

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class GeminiInvoiceProcessor(QThread):
    """Gemini发票处理线程"""
    
    # 定义信号
    progress_updated = pyqtSignal(int)
    processing_complete = pyqtSignal(list)
    processing_error = pyqtSignal(str)
    log_message = pyqtSignal(str)  # 添加日志信号
    
    def __init__(self, input_folder, excel_path, api_key, prompt):
        super().__init__()
        self.input_folder = input_folder
        self.excel_path = excel_path
        self.api_key = api_key
        self.prompt = prompt
        self.client = None
        
    def initialize_client(self):
        """初始化Gemini客户端"""
        try:
            log_msg = "初始化Gemini客户端"
            logger.info(log_msg)
            self.log_message.emit(log_msg)
            
            self.client = genai.Client(api_key=self.api_key)
            
            log_msg = "成功初始化Gemini客户端"
            logger.info(log_msg)
            self.log_message.emit(log_msg)
            return True
            
        except Exception as e:
            log_msg = f"初始化Gemini客户端失败: {str(e)}"
            logger.error(log_msg)
            self.log_message.emit(log_msg)
            return False
    
    def extract_invoice_data(self, file_path):
        """使用Gemini API提取发票数据"""
        try:
            logger.info(f"开始处理发票文件: {file_path}")
            
            # 初始化客户端
            if not self.client:
                if not self.initialize_client():
                    return None
            
            # 读取文件
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.pdf':
                # 处理PDF文件
                with open(file_path, 'rb') as f:
                    pdf_bytes = f.read()
                
                # 构建提示
                prompt = self.prompt
                
                # 调用Gemini API
                response = self.client.models.generate_content(
                    model="gemini-3-flash-preview",
                    contents=[
                        types.Part.from_bytes(
                            data=pdf_bytes,
                            mime_type='application/pdf',
                        ),
                        prompt
                    ],
                    config=types.GenerateContentConfig(
                        thinking_config=types.ThinkingConfig(thinking_budget=1024)
                    )
                )
                
                # 解析响应
                response_text = response.text
                logger.info(f"Gemini响应: {response_text}")
                
                # 尝试提取JSON - 支持单个对象或数组格式
                json_match = re.search(r'\[.*\]|\{.*\}', response_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group(0)
                    try:
                        # 清理JSON字符串中的控制字符，避免解析错误
                        json_str_clean = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', json_str)
                        # 尝试解析JSON
                        parsed_data = json.loads(json_str_clean)
                        
                        # 如果返回的是数组，返回所有发票对象
                        if isinstance(parsed_data, list) and len(parsed_data) > 0:
                            logger.info(f"检测到JSON数组格式，共{len(parsed_data)}个发票")
                            # 返回数组中的所有发票数据
                            return parsed_data
                        elif isinstance(parsed_data, dict):
                            # 单个对象，包装成数组返回
                            logger.info("检测到JSON对象格式")
                            return [parsed_data]
                        else:
                            logger.error(f"不支持的JSON格式: {type(parsed_data)}")
                            return None
                        
                        # 添加文件名
                        invoice_data['文件名'] = os.path.basename(file_path)
                        
                        # 验证关键字段是否存在，如果不存在则设置默认值
                        required_fields = [
                            '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                            '项目名称', '金额', '税率', '税额', '价税合计', '备注'
                        ]
                        
                        for field in required_fields:
                            if field not in invoice_data or invoice_data[field] is None:
                                invoice_data[field] = ''
                                logger.warning(f"字段 '{field}' 缺失或为空，设置为空字符串")
                        
                        # 记录成功提取的字段
                        logger.info(f"成功提取发票数据，发票号码: {invoice_data.get('发票号码', '未知')}")
                        return invoice_data
                        
                    except json.JSONDecodeError as e:
                        logger.error(f"无法解析JSON响应: {str(e)}")
                        logger.error(f"原始JSON字符串: {json_str}")
                        return None
                else:
                    logger.error("响应中未找到JSON格式数据")
                    logger.error(f"原始响应文本: {response_text}")
                    return None
                    
            elif file_ext in ['.png', '.jpg', '.jpeg']:
                # 处理图片文件
                with open(file_path, 'rb') as f:
                    image_bytes = f.read()
                
                # 确定MIME类型
                mime_type = 'image/jpeg' if file_ext in ['.jpg', '.jpeg'] else 'image/png'
                
                # 构建提示
                prompt = """
            
Role (角色设定):
你是一个专业的中国财务票据OCR智能解析助手。你的任务是从给定的发票文本/图片中，精准提取指定的11个关键字段。
Extraction Rules (核心提取逻辑 - 必须严格遵守):
发票号码 (Invoice Number):
⚠️ 警告：不要与左上角的“发票代码（Invoice Code，通常10或12位）”混淆。
逻辑：发票号码通常位于右上角，是 8位 数字（老版）或 20位 数字（全电发票）。请优先提取右上角标有“No.”或“发票号码”字样后的数字。
发票日期 (Date):
将日期统一标准化为 YYYYMMDD 格式或保持原始格式。
发票类型 (Type):
根据标题判断：
含“专用发票” -> 输出“增值税专用发票”
含“普通发票” -> 输出“增值税普通发票”
含“电子发票” -> 输出“电子发票”
若无明确标题，检查是否含“校验码”，有则通常为普通发票或专票。
购买方 vs 销售方 (Buyer vs Seller):
购买方名称：通常位于票据上方区域，关键字为“称：”或“名称：”。
销售方名称：通常位于票据下方区域。⚠️ 注意：销售方名称往往被红色的“发票专用章”覆盖，如果OCR识别出的文字包含乱码或不完整，请结合上下文或印章内容进行修正。
金额字段区分:
金额 = 不含税金额（Net Amount）。通常在表格右侧，列头为“金额”。
税额 = Tax Amount。
价税合计 = Total Amount。通常在发票底部，有“价税合计（大写）”和“（小写）”，请提取小写数字。
项目名称 (Items):
提取发票主要货物或应税劳务名称。如果是多行商品，请提取第一行或者概括主要内容。
备注：提取类似与2025.10-2025.11 服务费，之类的区间费用以及费用类型
Output Format (输出格式):
请直接输出合法的 JSON 格式，不要包含 Markdown 代码块以外的废话。如果某个字段在图上完全不存在，请填 null。
重要：金额、税额、价税合计可以是字符串格式（如"291262.14"），税率可以是字符串格式（如"3%"），这样更灵活且符合实际发票显示格式。
code
请按照要求输出以下JSON
{
  "发票号码": "String (确保不是发票代码)",
  "发票日期": "String",
  "发票类型": "String",
  "购买方名称": "String",
  "销售方名称": "String (注意印章干扰)",
  "项目名称": "String (货物或服务名称)",
  "金额": "String 或 Number (不含税金额)",
  "税率": "String 或 Number (如 0.06, 0.13 或 '6%', '13%')",
  "税额": "String 或 Number",
  "价税合计": "String 或 Number (最终支付金额)",
  "备注": "String (若无则填 null)"
}
                """
                
                # 调用Gemini API
                response = self.client.models.generate_content(
                    model="gemini-3-flash-preview",
                    contents=[
                        types.Part.from_bytes(
                            data=image_bytes,
                            mime_type=mime_type,
                        ),
                        prompt
                    ],
                    config=types.GenerateContentConfig(
                        thinking_config=types.ThinkingConfig(thinking_budget=1024)
                    )
                )
                
                # 解析响应
                response_text = response.text
                logger.info(f"Gemini响应: {response_text}")
                
                # 尝试提取JSON - 支持单个对象或数组格式
                json_match = re.search(r'\[.*\]|\{.*\}', response_text, re.DOTALL)
                if json_match:
                    json_str = json_match.group(0)
                    try:
                        # 清理JSON字符串中的控制字符，避免解析错误
                        json_str_clean = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', json_str)
                        
                        # 尝试解析JSON
                        parsed_data = json.loads(json_str_clean)
                        
                        # 如果返回的是数组，返回所有发票对象
                        if isinstance(parsed_data, list) and len(parsed_data) > 0:
                            logger.info(f"检测到JSON数组格式，共{len(parsed_data)}个发票")
                            # 处理数组中的每个发票，添加文件名
                            invoice_list = []
                            for invoice_data in parsed_data:
                                invoice_data['文件名'] = os.path.basename(file_path)
                                
                                # 验证关键字段是否存在，如果不存在则设置默认值
                                required_fields = [
                                    '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                                    '项目名称', '金额', '税率', '税额', '价税合计', '备注'
                                ]
                                
                                for field in required_fields:
                                    if field not in invoice_data or invoice_data[field] is None:
                                        invoice_data[field] = ''
                                        logger.warning(f"字段 '{field}' 缺失或为空，设置为空字符串")
                                
                                invoice_list.append(invoice_data)
                            
                            logger.info(f"成功提取{len(invoice_list)}个发票数据")
                            return invoice_list
                            
                        elif isinstance(parsed_data, dict):
                            # 单个对象，包装成数组返回
                            logger.info("检测到JSON对象格式")
                            parsed_data['文件名'] = os.path.basename(file_path)
                            
                            # 验证关键字段是否存在，如果不存在则设置默认值
                            required_fields = [
                                '发票号码', '发票日期', '发票类型', '购买方名称', '销售方名称',
                                '项目名称', '金额', '税率', '税额', '价税合计', '备注'
                            ]
                            
                            for field in required_fields:
                                if field not in parsed_data or parsed_data[field] is None:
                                    parsed_data[field] = ''
                                    logger.warning(f"字段 '{field}' 缺失或为空，设置为空字符串")
                            
                            logger.info(f"成功提取发票数据，发票号码: {parsed_data.get('发票号码', '未知')}")
                            return [parsed_data]
                        else:
                            logger.error(f"不支持的JSON格式: {type(parsed_data)}")
                            return None
                        
                    except json.JSONDecodeError as e:
                        logger.error(f"无法解析JSON响应: {str(e)}")
                        logger.error(f"原始JSON字符串: {json_str}")
                        return None
                else:
                    logger.error("响应中未找到JSON格式数据")
                    logger.error(f"原始响应文本: {response_text}")
                    return None
            else:
                logger.error(f"不支持的文件格式: {file_ext}")
                return None
                
        except Exception as e:
            logger.error(f"处理发票文件异常: {str(e)}")
            return None
    
    def run(self):
        """运行处理流程"""
        try:
            log_msg = "开始Gemini发票处理流程"
            logger.info(log_msg)
            self.log_message.emit(log_msg)
            
            # 获取文件列表
            files = []
            for file in os.listdir(self.input_folder):
                file_path = os.path.join(self.input_folder, file)
                if os.path.isfile(file_path) and file.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg')):
                    files.append(file_path)
            
            if not files:
                self.processing_error.emit("未找到支持的发票文件（PDF、PNG、JPG）")
                return
            
            log_msg = f"找到 {len(files)} 个发票文件"
            logger.info(log_msg)
            self.log_message.emit(log_msg)
            
            # 初始化客户端
            if not self.initialize_client():
                self.processing_error.emit("初始化Gemini客户端失败")
                return
            
            # 处理每个文件
            results = []
            total_files = len(files)
            
            for i, file_path in enumerate(files):
                log_msg = f"处理文件 {i+1}/{total_files}: {os.path.basename(file_path)}"
                logger.info(log_msg)
                self.log_message.emit(log_msg)
                
                # 提取发票数据
                invoice_data_list = self.extract_invoice_data(file_path)
                
                if invoice_data_list and isinstance(invoice_data_list, list):
                    # 成功提取到发票数组，添加所有发票
                    results.extend(invoice_data_list)
                    log_msg = f"成功提取{len(invoice_data_list)}个发票数据: {os.path.basename(file_path)}"
                    logger.info(log_msg)
                    self.log_message.emit(log_msg)
                else:
                    # 提取失败，添加一个失败记录
                    log_msg = f"提取发票数据失败: {os.path.basename(file_path)}"
                    logger.warning(log_msg)
                    self.log_message.emit(log_msg)
                    results.append({
                        '发票号码': '提取失败',
                        '发票日期': '提取失败',
                        '发票类型': '提取失败',
                        '购买方名称': '提取失败',
                        '销售方名称': '提取失败',
                        '项目名称': '提取失败',
                        '金额': '提取失败',
                        '税率': '提取失败',
                        '税额': '提取失败',
                        '价税合计': '提取失败',
                        '备注': '提取失败',
                        '文件名': os.path.basename(file_path)
                    })
                
                # 更新进度
                progress = int((i + 1) / total_files * 100)
                self.progress_updated.emit(progress)
            
            success_count = len([r for r in results if r['发票号码'] != '提取失败'])
            log_msg = f"处理完成，成功提取 {success_count}/{len(results)} 个发票"
            logger.info(log_msg)
            self.log_message.emit(log_msg)
            self.processing_complete.emit(results)
            
        except Exception as e:
            log_msg = f"处理流程异常: {str(e)}"
            logger.error(log_msg)
            self.log_message.emit(log_msg)
            self.processing_error.emit(f"处理流程异常: {str(e)}")


class GeminiInvoiceExtractorWidget(QWidget):
    """Gemini发票提取工具主界面"""
    
    def __init__(self):
        super().__init__()
        self.processor = None
        self.results = []
        self.init_ui()
        self.load_settings()
    
    def init_ui(self):
        """初始化UI界面"""
        layout = QVBoxLayout(self)
        
        # 创建分割器
        splitter = QSplitter(Qt.Vertical)
        
        # 上半部分：控制面板
        control_panel = QWidget()
        control_layout = QVBoxLayout(control_panel)
        
        # API Key输入
        api_group = QGroupBox("Gemini API配置")
        api_layout = QVBoxLayout()
        
        # API Key
        api_key_layout = QHBoxLayout()
        api_key_layout.addWidget(QLabel("API Key:"))
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("请输入Gemini API Key...")
        api_key_layout.addWidget(self.api_key_input)
        api_layout.addLayout(api_key_layout)
        
        # Secret Key (占位，保持与OCR工具布局一致)
        secret_key_layout = QHBoxLayout()
        secret_key_layout.addWidget(QLabel("Secret Key:"))
        self.secret_key_input = QLineEdit()
        self.secret_key_input.setPlaceholderText("Gemini API不需要Secret Key...")
        self.secret_key_input.setEnabled(False)
        secret_key_layout.addWidget(self.secret_key_input)
        api_layout.addLayout(secret_key_layout)
        
        api_group.setLayout(api_layout)
        control_layout.addWidget(api_group)
        
        # PDF文件夹选择
        pdf_group = QGroupBox("PDF文件夹")
        pdf_layout = QHBoxLayout()
        self.pdf_path_input = QLineEdit()
        self.pdf_path_input.setPlaceholderText("选择包含PDF发票的文件夹...")
        pdf_browse_btn = QPushButton("浏览...")
        pdf_browse_btn.clicked.connect(self.browse_pdf_folder)
        pdf_layout.addWidget(self.pdf_path_input)
        pdf_layout.addWidget(pdf_browse_btn)
        pdf_group.setLayout(pdf_layout)
        control_layout.addWidget(pdf_group)
        
        # Excel文件选择
        excel_group = QGroupBox("Excel输出文件")
        excel_layout = QHBoxLayout()
        self.excel_path_input = QLineEdit()
        self.excel_path_input.setPlaceholderText("选择或创建Excel文件...")
        excel_browse_btn = QPushButton("浏览...")
        excel_browse_btn.clicked.connect(self.browse_excel_file)
        excel_layout.addWidget(self.excel_path_input)
        excel_layout.addWidget(excel_browse_btn)
        excel_group.setLayout(excel_layout)
        control_layout.addWidget(excel_group)
        
        # 提示词配置
        prompt_group = QGroupBox("提示词配置")
        prompt_layout = QVBoxLayout()
        self.prompt_text = QTextEdit()
        self.prompt_text.setMaximumHeight(200)
        self.prompt_text.setPlaceholderText("输入用于PDF发票提取的提示词...")
        self.prompt_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
                font-family: monospace;
                font-size: 9pt;
            }
        """)
        prompt_layout.addWidget(self.prompt_text)
        prompt_group.setLayout(prompt_layout)
        control_layout.addWidget(prompt_group)
        
        # 处理控制
        process_group = QGroupBox("处理控制")
        process_layout = QVBoxLayout()
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        process_layout.addWidget(self.progress_bar)
        
        # 添加日志框
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(150)
        self.log_text.setStyleSheet("""
            QTextEdit {
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: #f9f9f9;
                font-family: monospace;
                font-size: 9pt;
            }
        """)
        process_layout.addWidget(self.log_text)
        
        # 处理按钮
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始提取")
        self.start_btn.clicked.connect(self.start_processing)
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 12px 24px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #3d8b40;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        self.stop_btn = QPushButton("停止")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-weight: bold;
                padding: 12px 24px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
            QPushButton:pressed {
                background-color: #c62828;
            }
        """)
        
        self.save_btn = QPushButton("保存结果")
        self.save_btn.clicked.connect(self.save_results)
        self.save_btn.setEnabled(False)
        
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.stop_btn)
        btn_layout.addWidget(self.save_btn)
        btn_layout.addStretch()
        
        process_layout.addLayout(btn_layout)
        process_group.setLayout(process_layout)
        control_layout.addWidget(process_group)
        
        # 统计信息
        self.stats_label = QLabel("准备就绪")
        self.stats_label.setStyleSheet("font-weight: bold; color: #666;")
        control_layout.addWidget(self.stats_label)
        
        control_layout.addStretch()
        
        # 下半部分：结果显示
        result_widget = QWidget()
        result_layout = QVBoxLayout(result_widget)
        
        # 结果表格
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(12)
        self.result_table.setHorizontalHeaderLabels([
            "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
            "项目名称", "金额", "税率", "税额", "价税合计", "备注", "文件名"
        ])
        
        # 设置表格样式
        self.result_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #e3f2fd;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: 1px solid #ddd;
                font-weight: bold;
            }
        """)
        
        # 设置列宽
        self.result_table.setColumnWidth(0, 150)  # 发票号码
        self.result_table.setColumnWidth(1, 100)  # 发票日期
        self.result_table.setColumnWidth(2, 80)   # 发票类型
        self.result_table.setColumnWidth(3, 200)  # 购买方名称
        self.result_table.setColumnWidth(4, 200)  # 销售方名称
        self.result_table.setColumnWidth(5, 200)  # 项目名称
        self.result_table.setColumnWidth(6, 100)  # 金额
        self.result_table.setColumnWidth(7, 80)   # 税率
        self.result_table.setColumnWidth(8, 100)  # 税额
        self.result_table.setColumnWidth(9, 100)  # 价税合计
        self.result_table.setColumnWidth(10, 200) # 备注
        self.result_table.setColumnWidth(11, 150) # 文件名
        
        result_layout.addWidget(self.result_table)
        
        # 添加到分割器
        splitter.addWidget(control_panel)
        splitter.addWidget(result_widget)
        
        # 设置分割器比例
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        
        layout.addWidget(splitter)
    
    def browse_pdf_folder(self):
        """浏览PDF文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择PDF文件夹", self.pdf_path_input.text())
        if folder:
            self.pdf_path_input.setText(folder)
    
    def browse_excel_file(self):
        """浏览Excel文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "选择Excel文件", self.excel_path_input.text(), "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            if not file_path.endswith(('.xlsx', '.xls')):
                file_path += '.xlsx'
            self.excel_path_input.setText(file_path)
    
    def start_processing(self):
        """开始处理"""
        # 检查是否已有处理线程在运行
        if self.processor and self.processor.isRunning():
            QMessageBox.warning(self, "警告", "已有处理任务在进行中，请等待当前任务完成")
            return
            
        pdf_folder = self.pdf_path_input.text()
        excel_path = self.excel_path_input.text()
        api_key = self.api_key_input.text()
        
        if not pdf_folder:
            QMessageBox.warning(self, "警告", "请选择PDF文件夹")
            return
        
        if not excel_path:
            QMessageBox.warning(self, "警告", "请选择Excel输出文件")
            return
        
        if not api_key:
            QMessageBox.warning(self, "警告", "请输入API Key")
            return
        
        # 检查文件夹是否存在
        if not os.path.exists(pdf_folder):
            QMessageBox.warning(self, "警告", f"PDF文件夹不存在: {pdf_folder}")
            return
            
        # 检查文件夹中是否有支持的文件
        files = [f for f in os.listdir(pdf_folder) if f.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg'))]
        if not files:
            QMessageBox.warning(self, "警告", f"文件夹中没有找到支持的发票文件（PDF、PNG、JPG）")
            return
        
        # 清空结果表格和日志
        self.result_table.setRowCount(0)
        self.log_text.clear()
        
        # 创建处理器
        self.processor = GeminiInvoiceProcessor(pdf_folder, excel_path, api_key, self.prompt_text.toPlainText())
        self.processor.progress_updated.connect(self.update_progress)
        self.processor.processing_complete.connect(self.display_results)
        self.processor.processing_error.connect(self.show_error)
        self.processor.log_message.connect(self.append_log)
        
        # 更新UI状态
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.save_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.stats_label.setText("正在初始化...")
        
        # 启动处理线程
        try:
            self.processor.start()
            self.stats_label.setText("处理中...")
        except Exception as e:
            logger.error(f"启动处理线程失败: {str(e)}")
            self.show_error(f"启动处理失败: {str(e)}")
            self.reset_ui_state()
    
    def stop_processing(self):
        """停止处理"""
        if self.processor and self.processor.isRunning():
            self.processor.terminate()
            self.processor.wait()
            
            # 更新UI状态
            self.start_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            self.save_btn.setEnabled(True)
            self.progress_bar.setVisible(False)
            self.stats_label.setText("处理已停止")
    
    def update_progress(self, value):
        """更新进度"""
        self.progress_bar.setValue(value)
        self.stats_label.setText(f"处理进度: {value}%")
    
    def display_results(self, results):
        """显示结果"""
        self.results = results
        
        # 验证结果数据
        if not results:
            QMessageBox.warning(self, "警告", "没有提取到任何发票数据")
            self.reset_ui_state()
            return
        
        # 更新表格
        self.result_table.setRowCount(len(results))
        
        for row, data in enumerate(results):
            # 验证数据格式
            if not isinstance(data, dict):
                logger.error(f"第 {row+1} 行数据格式错误: 期望dict，得到 {type(data)}")
                # 创建错误行
                for col in range(12):
                    item = QTableWidgetItem("数据格式错误")
                    item.setBackground(Qt.red)
                    self.result_table.setItem(row, col, item)
                continue
            
            # 填充表格数据
            for col, key in enumerate([
                "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
                "项目名称", "金额", "税率", "税额", "价税合计", "备注", "文件名"
            ]):
                try:
                    value = data.get(key, "")
                    # 确保值为字符串
                    if value is None:
                        value = ""
                    else:
                        value = str(value)
                    
                    item = QTableWidgetItem(value)
                    
                    # 标记提取失败的行
                    if key == "发票号码" and value == "提取失败":
                        # 整行标记为失败
                        for fail_col in range(12):
                            fail_item = QTableWidgetItem(str(data.get([
                                "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
                                "项目名称", "金额", "税率", "税额", "价税合计", "备注", "文件名"
                            ][fail_col], "")))
                            fail_item.setBackground(Qt.lightGray)
                            self.result_table.setItem(row, fail_col, fail_item)
                        break
                    
                    self.result_table.setItem(row, col, item)
                    
                except Exception as e:
                    logger.error(f"填充表格时出错 (行{row+1}, 列{col}, 字段{key}): {str(e)}")
                    item = QTableWidgetItem("数据错误")
                    item.setBackground(Qt.red)
                    self.result_table.setItem(row, col, item)
        
        # 更新UI状态
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.save_btn.setEnabled(True)
        self.progress_bar.setVisible(False)
        
        # 统计信息
        success_count = len([r for r in results if isinstance(r, dict) and r.get('发票号码', '') != '提取失败'])
        total_count = len(results)
        self.stats_label.setText(f"处理完成: {success_count}/{total_count} 个发票成功提取")
        
        # 记录详细统计
        logger.info(f"结果显示完成 - 总计: {total_count}, 成功: {success_count}, 失败: {total_count - success_count}")
        
        # 保存设置
        self.save_settings()
    
    def reset_ui_state(self):
        """重置UI状态"""
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.save_btn.setEnabled(len(self.results) > 0)
        self.progress_bar.setVisible(False)
        self.stats_label.setText("准备就绪")
    
    def append_log(self, message):
        """添加日志到日志框"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
    
    def show_error(self, message):
        """显示错误"""
        # 对于客户端初始化失败，不显示错误框，只记录日志
        if "初始化Gemini客户端失败" in message:
            self.append_log(f"错误: {message}")
            self.reset_ui_state()
        else:
            QMessageBox.critical(self, "错误", message)
            # 重置UI状态
            self.reset_ui_state()
    
    def save_results(self):
        """保存结果到Excel"""
        if not self.results:
            QMessageBox.warning(self, "警告", "没有可保存的结果")
            return
        
        excel_path = self.excel_path_input.text()
        if not excel_path:
            QMessageBox.warning(self, "警告", "请选择Excel输出文件")
            return
        
        try:
            # 检查文件是否存在
            if os.path.exists(excel_path):
                # 加载现有工作簿
                wb = load_workbook(excel_path)
                ws = wb.active
                # 找到最后一行
                last_row = ws.max_row + 1
            else:
                # 创建新工作簿
                wb = Workbook()
                ws = wb.active
                # 添加标题行
                headers = [
                    "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
                    "项目名称", "金额", "税率", "税额", "价税合计", "备注", "文件名"
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    # 设置标题样式
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                last_row = 2
            
            # 添加数据
            for data in self.results:
                for col, key in enumerate([
                    "发票号码", "发票日期", "发票类型", "购买方名称", "销售方名称",
                    "项目名称", "金额", "税率", "税额", "价税合计", "备注", "文件名"
                ], 1):
                    value = data.get(key, "")
                    ws.cell(row=last_row, column=col, value=value)
                    # 设置单元格样式
                    ws.cell(row=last_row, column=col).alignment = Alignment(horizontal='left', vertical='center')
                last_row += 1
            
            # 保存工作簿
            wb.save(excel_path)
            
            QMessageBox.information(self, "成功", f"结果已保存到: {excel_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存Excel文件失败: {str(e)}")
            logger.error(f"保存Excel文件失败: {str(e)}")
    
    def load_settings(self):
        """加载设置"""
        try:
            if os.path.exists('toolbox_config.json'):
                with open('toolbox_config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                if 'gemini_invoice_extractor' in config:
                    gemini_config = config['gemini_invoice_extractor']
                    self.pdf_path_input.setText(gemini_config.get('input_folder', ''))
                    self.excel_path_input.setText(gemini_config.get('excel_file', ''))
                    self.api_key_input.setText(gemini_config.get('api_key', ''))
                    prompt = gemini_config.get('prompt', '')
                    if not prompt:
                        prompt = """
请从这张发票PDF中提取以下信息，并以JSON格式返回，允许金额、税额、价税合计、税率为字符串或数字格式：
 
{ 
 
"发票号码": "", 
 
"发票日期": "", 
 
"发票类型": "", 
 
"购买方名称": "", 
 
"销售方名称": "", 
 
"项目名称": "", 
 
"金额": "", 
 
"税率": "", 
 
"税额": "", 
 
"价税合计": "", 
 
"备注": "" 
 
} 
 
重要提示：
- 金额、税额、价税合计可以是字符串（如"291262.14"）或数字格式
- 税率可以是字符串（如"3%"）或数字格式（如0.03）
- 这样可以更灵活地处理不同格式的发票数据
- 把民航基金的金额输出到备注
如果有多页PDF，请精准定位到发票页（没有各种语言的"发票"字样的都不是发票页，各种语言的发票你都需要识别）。可能会有很多张发票，你每张发票都需要输出上述JSON。
                        """
                    self.prompt_text.setText(prompt)
        except Exception as e:
            logger.error(f"加载设置失败: {str(e)}")
    
    def save_settings(self):
        """保存设置"""
        try:
            # 加载现有配置
            config = {}
            if os.path.exists('toolbox_config.json'):
                with open('toolbox_config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            # 更新Gemini发票提取配置
            config['gemini_invoice_extractor'] = {
                'input_folder': self.pdf_path_input.text(),
                'excel_file': self.excel_path_input.text(),
                'api_key': self.api_key_input.text(),
                'prompt': self.prompt_text.toPlainText()
            }
            
            # 保存配置
            with open('toolbox_config.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
                
        except Exception as e:
            logger.error(f"保存设置失败: {str(e)}")